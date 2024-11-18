import openpyxl
from openpyxl import load_workbook
from typing import Generator
from rigour.mime.types import XLSX
from normality import stringify, slugify
from datetime import datetime

from zavod import Context, helpers as h
from zavod.logic.pep import categorise

RESOURCES = [
    ("dk", "dan", "dk.xlsx", "PEP_listen.xlsx"),
    ("fo", "fao", "fo.xlsx", "PEP_listen f"),
    ("gl", "kal", "gl.xlsx", "PEP_listen g"),
]
CURRENT_HEADERS = [
    None,
    "Navn",
    None,
    "Stillingsbetegnelse ",
    "Fødselsdato",
    "Tilføjet på PEP-listen (dato)",
    "Ny på PEP-listen (x)",
]


def header_names(cells):
    headers = []
    for idx, cell in enumerate(cells):
        if cell is None:
            cell = f"column_{idx}"
        headers.append(slugify(cell, "_").lower())
    return headers


def parse_old_pep(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
) -> Generator[dict, None, None]:
    """This function parses the old PEP list, which is in the format of a table in the second sheet of each Excel file.

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The sheet to parse.

    Yields:
        Generator[dict, None, None]: Each row of the table as a dictionary.
    """

    headers = None
    for idx, row in enumerate(sheet.rows):
        cells = [c.value for c in row]
        if headers is None:
            headers = header_names(cells)
            continue
        record = {}
        for header, value in zip(headers, cells):
            if isinstance(value, datetime):
                value = value.date()
            value = stringify(value)
            if value is not None:
                record[header] = value
        if len(record) == 0:
            continue
        yield record


def parse_current_pep(
    sheet: openpyxl.worksheet.worksheet.Worksheet, context: Context
) -> Generator[dict, None, None]:
    """This function parses the current PEP list, which is in the format of multiple tables in the first sheet of the Excel file.

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The sheet to parse.

    Yields:
        Generator[dict, None, None]: Each row of the table as a dictionary.
    """
    for idx, row in enumerate(sheet.rows):
        # First row is just the date and name of the dataset
        # Second row is the header
        if idx == 0:
            continue
        elif idx == 1:
            heads = [c.value for c in row]
            assert heads[:7] == CURRENT_HEADERS, heads[:7]
            continue

        # Every row can be either:
        # - Name of the organization
        # - Data of the PEP

        # If there are no values in the row, then it's an empty row
        if all([c.value is None for c in row]):
            pass

        # If there is only one not null value in the row and it's the first, then it's the name of the list
        elif sum([c.value is not None for c in row]) == 1 and row[0].value is not None:
            list_name = row[0].value

        # If at least one of the 2nd, 3rd, 4th, 5th and 6th columns have values, then it's a PEP data
        elif any([c.value is not None for c in row[1:6]]):
            # Meaning "No boards"
            if row[1].value == "Ingen styrelser etc.":
                continue

            yield {
                "last-name": row[1].value,
                "first-name": row[2].value,
                "position": row[3].value,
                "birth-date": row[4].value,
                "start-date": row[5].value,
                "list-name": list_name,
            }

        else:
            context.log.warning(f"Couldn't parse row: {[c.value for c in row]}")


def crawl_current_pep_item(context: Context, country: str, lang: str, input_dict: dict):
    first_name = input_dict.pop("first-name")
    last_name = input_dict.pop("last-name")

    entity = context.make("Person")
    entity.id = context.make_slug(first_name, last_name)
    h.apply_name(entity, first_name=first_name, last_name=last_name)

    birth_date = input_dict.pop("birth-date")
    h.apply_date(entity, "birthDate", birth_date.strip() if birth_date else None)
    list_name = input_dict.pop("list-name")
    affiliation = input_dict.pop("position")

    if list_name == "Medlemmer af Folketinget":
        position_name = "Member of the Folketing"
    elif list_name == "Medlemmer af Europaparlamentet":
        position_name = "Member of the European Parliament"
    elif "Medlemmer af Lagtinget" in list_name:
        position_name = "Member of the Lagting"
    elif "Medlemmer" in list_name and "Inatsisartut" in list_name:
        position_name = "Member of the Inatsisartut"
    elif "partiernes styrelsesorganer" in list_name:
        position_name = "Member of leadership of " + (
            affiliation or "a political party"
        )
    else:
        position_name = affiliation
    assert position_name != "Socialdemokratiet"
    assert position_name is not None, entity.id

    position = h.make_position(context, position_name, country=country, lang=lang)
    categorisation = categorise(context, position, is_pep=True)

    occupancy = h.make_occupancy(
        context,
        entity,
        position,
        True,
        start_date=input_dict.pop("start-date").strip(),
        categorisation=categorisation,
    )

    if occupancy:
        context.emit(entity, target=True)
        context.emit(position)
        context.emit(occupancy)

    context.audit_data(input_dict)


def crawl_old_pep_item(context: Context, country: str, lang: str, input_dict: dict):
    last_name = input_dict.pop("efternavn")
    first_name = input_dict.pop("fornavn")

    position_col = "stilling" if "stilling" in input_dict else "stillingsbetegnelse"

    entity = context.make("Person")
    entity.id = context.make_slug(first_name, last_name)
    h.apply_name(entity, first_name=first_name, last_name=last_name)

    if "fodselsdato" in input_dict:
        h.apply_date(entity, "birthDate", input_dict.pop("fodselsdato").strip())

    position = h.make_position(
        context, input_dict.pop(position_col), country=country, lang=lang
    )

    occupation = h.make_occupancy(
        context,
        entity,
        position,
        True,
        end_date=input_dict.pop("fjernet_fra_pep_listen_dato").strip(),
        categorisation=categorise(context, position, is_pep=True),
    )

    if occupation:
        context.emit(entity, target=True)
        context.emit(position)
        context.emit(occupation)

    context.audit_data(input_dict)


def crawl(context: Context):
    doc = context.fetch_html(context.data_url)
    doc.make_links_absolute(context.data_url)

    for country, lang, name, file_pattern in RESOURCES:
        links = doc.xpath(f'//a[contains(@href, "{file_pattern}")]')
        assert len(links) == 1, (file_pattern, links)
        path = context.fetch_resource(name, links[0].get("href"))
        context.export_resource(path, XLSX, title=context.SOURCE_TITLE)

        wb = load_workbook(path, read_only=True)

        # Crawl old PEP list
        old_sheet = (
            "Tidligere PEP'ere"
            if "Tidligere PEP'ere" in wb.sheetnames
            else "Tidligere PEP'er"
        )
        for item in parse_old_pep(wb[old_sheet]):
            crawl_old_pep_item(context, country, lang, item)

        new_sheet = (
            "Nuværende PEP'ere"
            if "Nuværende PEP'ere" in wb.sheetnames
            else "Nuværende PEP'er"
        )
        # Crawl current PEP list
        for item in parse_current_pep(wb[new_sheet], context):
            crawl_current_pep_item(context, country, lang, item)
